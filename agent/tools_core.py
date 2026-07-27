"""
tools_core.py — 数据操作工具
read_data_file / query_loaded_data / clean_data / undo / export_data_file / merge_file
"""
import logging
import shutil
from pathlib import Path

import numpy as np
import pandas as pd

from langchain.tools import tool as lc_tool

from core import config
from core.context import DataFrameContext
from core.file_io import (
    probe_csv_encoding, read_csv_full, read_csv_sample, read_csv_chunked,
    df_meta, build_summary, get_file_size_mb,
)
from core.cleaning import (
    is_aggregation, eval_chunked, eval_full_reload,
    clean_in_memory, squash_large,
)

_log = logging.getLogger(__name__)

SAMPLE_ROWS = 100  # read_data_file 采样行数


# ── 共享辅助 ──────────────────────────────────────────────

def _file_to_slot_name(stem: str) -> str:
    """将文件名映射到简短的槽位名。去掉扩展名和特殊字符，限制长度。"""
    # 去掉常见的日期/版本后缀，取核心名称
    import re
    cleaned = re.sub(r'[-_](截至|截止|每日|汇总|数据|表).*$', '', stem)
    cleaned = re.sub(r'[^\w一-鿿]', '_', cleaned)
    return cleaned[:12] if cleaned else stem[:12]


def _resolve_df(ctx: DataFrameContext) -> pd.DataFrame | None:
    """获取当前上下文中可用的 DataFrame，采样模式下自动触发全量加载。"""
    if ctx.loaded_df is not None:
        return ctx.loaded_df
    if ctx.source_file_path:
        _ensure_full_loaded(ctx)
        return ctx.loaded_df
    if ctx.large_file_path:
        try:
            df, _ = read_csv_full(ctx.large_file_path)
            return df
        except Exception:
            return None
    return None


def _resolve_read_path(file_path: str, ctx: DataFrameContext | None = None) -> str:
    """解析读取路径。绝对路径直接返回；相对路径先查 raw/，再查会话 workspace，最后全局 workspace。
    自动处理前导 raw/ 或 workspace/ 前缀（Agent 可能从 list_files 输出中提取）。"""
    raw = config.RAW_DIR.resolve()
    global_ws = config.WORKSPACE_DIR.resolve()
    session_ws = Path(ctx.workspace_dir).resolve() if ctx else global_ws

    # 去掉前导 raw/ 和 workspace/（避免 raw/raw/ 双重拼接）
    cleaned = file_path
    for prefix in ("raw/", "raw\\", "workspace/", "workspace\\"):
        if cleaned.lower().startswith(prefix.lower()):
            cleaned = cleaned[len(prefix):]
            break

    path = Path(cleaned)
    if path.is_absolute():
        resolved = str(path.resolve())
        if (resolved.startswith(str(raw))
            or resolved.startswith(str(session_ws))
            or resolved.startswith(str(global_ws))):
            return resolved
        if Path(resolved).exists():
            return resolved
        return f"错误: 文件不存在 — {resolved}"
    # 搜索顺序: raw → 会话 workspace → 全局 workspace
    for base in (raw, session_ws, global_ws):
        if base == global_ws and base == session_ws:
            continue  # 避免重复检查
        candidate = (base / path).resolve()
        if candidate.exists() and (str(candidate).startswith(str(base))
                                   or str(candidate).startswith(str(global_ws))):
            return str(candidate)
    return str((raw / path).resolve())


def _resolve_write_path(file_path: str, ctx: DataFrameContext | None = None) -> tuple[str, str]:
    """解析写入路径。只允许写入会话 workspace（或全局 workspace 回退）。返回 (abs_path, error_msg)。"""
    raw = config.RAW_DIR.resolve()
    global_ws = config.WORKSPACE_DIR.resolve()
    ws = Path(ctx.workspace_dir).resolve() if ctx else global_ws

    # 去掉前导 raw/ 和 workspace/（Agent 常加前缀，避免双重拼接）
    cleaned = file_path
    for prefix in ("raw/", "raw\\", "workspace/", "workspace\\"):
        if cleaned.lower().startswith(prefix.lower()):
            cleaned = cleaned[len(prefix):]
            break

    path = Path(cleaned)
    if path.is_absolute():
        resolved = str(path.resolve())
        if resolved.startswith(str(raw)):
            return "", "错误: raw/ 为原始数据区，禁止写入。请保存到 workspace/ 目录。"
        if resolved.startswith(str(ws)) or resolved.startswith(str(global_ws)):
            return resolved, ""
        return "", f"错误: 只能写入 workspace/ 目录。当前路径: {resolved}"
    return str((ws / path).resolve()), ""


# ── 延迟全量加载 ──────────────────────────────────────────

def _ensure_full_loaded(ctx: DataFrameContext) -> bool:
    """确保数据已全量加载到 ctx.loaded_df。
    仅当 ctx 处于采样模式（source_file_path 非空）时触发加载。
    大文件模式（large_file_path）不触发，走分块查询。
    返回 True 表示 loaded_df 可用。"""
    if ctx.loaded_df is not None:
        return True
    if not ctx.source_file_path:
        return False  # 大文件模式或不曾加载，不处理

    suffix = ctx.source_file_suffix
    _log.info("延迟全量加载: %s", ctx.source_file_path)
    try:
        if suffix == ".csv":
            df, _enc = read_csv_full(ctx.source_file_path)
        elif suffix in (".xlsx", ".xls"):
            df = pd.read_excel(ctx.source_file_path)
        else:
            return False

        ctx.loaded_df = df
        ctx.loaded_filename = Path(ctx.source_file_path).name
        # 清除源文件追踪和大文件标记
        ctx.source_file_path = ""
        ctx.source_file_suffix = ""
        ctx.large_file_path = ""
        ctx.large_file_encoding = ""
        ctx.large_file_suffix = ""
        _log.info("延迟加载完成: %s (%d 行)", ctx.loaded_filename, len(df))
        return True
    except Exception as e:
        _log.error("延迟加载失败: %s — %s", ctx.source_file_path, e)
        return False


# ── 工具工厂 ──────────────────────────────────────────────

def create_core_tools(ctx: DataFrameContext) -> list:
    """创建核心数据操作工具（9 个），仅依赖 ctx。"""

    # ── list_files ───────────────────────────────────────

    @lc_tool
    def list_files(directory: str = "") -> str:
        """
        列出 raw/ 或 workspace/ 目录下的文件（CSV/Excel），含文件大小。
        用于在 read_data_file 之前确认文件的确切名称和路径。

        Parameters:
        directory: 可选 'raw' 或 'workspace'，不传则两个都列
        """
        from core.config import RAW_DIR
        from pathlib import Path
        ws_dir = Path(ctx.workspace_dir)
        results = []
        for label, base in [("raw", RAW_DIR), ("workspace", ws_dir)]:
            if directory and directory.lower().strip("/\\") != label:
                continue
            if not base.exists():
                results.append(f"⚠️ {label}/ 目录不存在")
                continue
            files = sorted(base.glob("*"))
            if not files:
                results.append(f"📂 {label}/ (空)")
                continue
            lines = [f"📂 {label}/"]
            for f in files:
                if f.is_file():
                    suffix = f.suffix.lower()
                    if suffix in (".csv", ".xlsx", ".xls"):
                        size_kb = round(f.stat().st_size / 1024, 1)
                        marker = "🔶" if size_kb > 100 * 1024 else ""
                        lines.append(f"  {f.name}  ({size_kb:.0f} KB) {marker}")
                elif f.is_dir():
                    lines.append(f"  📁 {f.name}/")
            results.append("\n".join(lines))
        if not results:
            return "错误: 目录参数无效，请用 'raw' 或 'workspace'。"
        return "\n\n".join(results)

    # ── read_data_file ────────────────────────────────────

    @lc_tool
    def read_data_file(file_path: str) -> str:
        """
        抽样探索文件结构：只读前 100 行，返回列名/dtype/样本，不加载全量。
        后续首次 query/clean/export 时自动触发全量加载。
        Excel >100MB 拒绝并提示转 CSV；CSV >100MB 分块扫描仅缓存摘要。

        Parameters:
        file_path: 相对路径（raw/ 或 workspace/ 下）或用户提供的绝对路径
        """
        resolved = _resolve_read_path(file_path, ctx)
        if not resolved or resolved.startswith("错误:"):
            return resolved if resolved else (
                f"错误: 文件路径不受支持 — {file_path}\n"
                f"可用路径: raw/ 或 workspace/ 下的相对路径，或用户提供的绝对路径。\n"
                f"示例: read_data_file('sales.csv') 或 read_data_file('D:/data/report.xlsx')"
            )
        file_path = resolved

        path = Path(file_path)
        if not path.exists():
            return f"错误: 文件不存在 — {file_path}"

        # ★ 防覆盖保护：当前槽位有数据时，自动切到规范命名的槽位
        if ctx.loaded_df is not None or ctx.source_file_path:
            slot_name = _file_to_slot_name(path.stem)
            if slot_name != ctx.active_slot:
                _log.info("read_data_file 自动切槽: '%s' → '%s'", ctx.active_slot, slot_name)
                ctx.switch(slot_name)

        suffix = path.suffix.lower()
        if suffix not in (".csv", ".xlsx", ".xls"):
            return f"错误: 不支持的文件格式 {suffix}"

        original_name = path.name
        n_sheets = 0
        sheet_names = []

        # ── Excel: 采样前 N 行探查结构 ──
        if suffix in (".xlsx", ".xls"):
            file_size_mb = get_file_size_mb(file_path)
            if file_size_mb > config.LARGE_FILE_MB:
                return (
                    f"错误: Excel 文件过大 ({file_size_mb:.1f} MB > {config.LARGE_FILE_MB} MB 限制)\n"
                    f"请先在 Excel 中将文件另存为 CSV 格式，再用 read_data_file 读取该 CSV。\n"
                    f"CSV 大文件支持分块读取，不会全量加载到内存。"
                )

            # 多 sheet 检测（仅元数据，不加载数据）
            try:
                xl = pd.ExcelFile(file_path)
                sheet_names = xl.sheet_names
                n_sheets = len(sheet_names)
                if n_sheets > 1:
                    _log.info("多sheet Excel: %s (%d sheets)", path.name, n_sheets)
            except Exception:
                n_sheets = 1
                sheet_names = []

            # ★ 仅采样前 SAMPLE_ROWS 行
            try:
                sample_df = pd.read_excel(file_path, nrows=SAMPLE_ROWS)
            except Exception as e:
                return f"错误: 读取 Excel 失败 — {e}"

            sample_df.columns = [str(c) for c in sample_df.columns]
            ctx.loaded_df = None                     # 不驻留内存（Excel：采样模式）
            ctx.loaded_filename = original_name
            ctx.large_file_path = ""
            ctx.large_file_encoding = ""
            ctx.large_file_suffix = ""
            ctx.source_file_path = file_path         # ★ 记录源文件，首次查询时全量加载
            ctx.source_file_suffix = suffix
            _log.info("Excel 采样: %s (%d 采样行, %.1f MB)",
                      original_name, len(sample_df), file_size_mb)

            # .xlsx 快速取总行数（openpyxl read_only，仅读 XML 元数据）
            excel_rows_note = ""
            if suffix == ".xlsx":
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(file_path, read_only=True)
                    total = wb.active.max_row
                    wb.close()
                    if total is not None and total > SAMPLE_ROWS:
                        excel_rows_note = (
                            f"\n📐 Excel 总行数约 {total:,}"
                            f"（当前仅采样前 {SAMPLE_ROWS} 行探查）"
                        )
                except Exception:
                    pass

            meta = df_meta(sample_df)
            label = f"已采样（仅前{SAMPLE_ROWS}行）"
            result = build_summary(meta, original_name, "n/a", is_chunked=False,
                                   label=label)
            result += excel_rows_note
            result += (
                f"\n\n💡 仅加载前 {SAMPLE_ROWS} 行用于探查数据结构。"
                f"\n   首次执行 query / clean / export 时将自动全量加载。"
            )

            # 多 sheet 提示
            if n_sheets > 1:
                result = (
                    f"⚠️ 此 Excel 包含 {n_sheets} 个工作表: {', '.join(sheet_names[:8])}"
                    f"{'…' if n_sheets > 8 else ''}\n"
                    f"当前仅探查第一个 sheet ({sheet_names[0]})。\n"
                    f"如需加载其他 sheet，请先在 Excel 中将目标 sheet 另存为独立文件。\n\n"
                    + result
                )
            return result

        # ── CSV: 按大小选择探查策略 ──
        file_size_mb = get_file_size_mb(file_path)
        if file_size_mb > config.LARGE_FILE_MB:
            # 大 CSV：分块扫描获取准确元数据（不驻留内存）
            enc = probe_csv_encoding(file_path)
            try:
                meta = read_csv_chunked(file_path, encoding=enc)
            except Exception as e:
                return f"错误: 分块读取失败 — {e}"
            ctx.loaded_df = None
            ctx.loaded_filename = original_name
            ctx.large_file_path = file_path        # 大文件走分块查询
            ctx.large_file_encoding = enc
            ctx.large_file_suffix = ".csv"
            ctx.source_file_path = ""              # 大文件不延迟加载
            ctx.source_file_suffix = ""
            _log.info("大文件扫描: %s (%d 行)", original_name, meta["row_count"])
            result = build_summary(meta, original_name, enc, is_chunked=True)
        else:
            # 小 CSV：采样前 SAMPLE_ROWS 行 + 尝试读尾部探查总行数
            try:
                sample_df, detected_enc = read_csv_sample(file_path, nrows=SAMPLE_ROWS)
            except Exception as e:
                return f"错误: 读取文件失败 — {e}"

            # 快速统计总行数（缓冲读取换行符计数，不解析 CSV）
            total_rows = 0
            tail_note = ""
            try:
                with open(file_path, "rb") as f:
                    buf_size = 1024 * 1024  # 1 MB 缓冲
                    while True:
                        buf = f.read(buf_size)
                        if not buf:
                            break
                        total_rows += buf.count(b"\n")
                total_rows = max(0, total_rows - 1)  # 减去表头行
            except Exception:
                _log.warning("行数统计失败，跳过: %s", file_path)

            # 若估计值在阈值附近（≤200 行），用 csv.reader 精确计数
            # csv.reader 正确处理引号内嵌换行符，弥补 buf.count 的偏差
            if 0 < total_rows <= SAMPLE_ROWS * 2:
                try:
                    import csv
                    with open(file_path, "r", encoding=detected_enc,
                              newline="") as f:
                        reader = csv.reader(f)
                        accurate = sum(1 for _ in reader)
                    total_rows = max(0, accurate - 1)  # 减表头，精确值
                except Exception:
                    pass  # 保留换行计数的估计值

            if total_rows > 0:
                tail_note = (
                    f"\n📐 估计行数 ≈{total_rows:,}"
                    f"（当前仅采样前 {SAMPLE_ROWS} 行探查）"
                )

            # 极小文件（≤SAMPLE_ROWS 行）→ 直接全量驻留
            if 0 < total_rows <= SAMPLE_ROWS:
                try:
                    full_df, detected_enc2 = read_csv_full(file_path)
                    full_df.columns = [str(c).lstrip("﻿") for c in full_df.columns]
                    ctx.loaded_df = full_df
                    ctx.loaded_filename = original_name
                    ctx.large_file_path = ""
                    ctx.large_file_encoding = ""
                    ctx.large_file_suffix = ""
                    ctx.source_file_path = ""
                    ctx.source_file_suffix = ""
                    _log.info("小文件直接加载: %s (%d 行)", original_name, len(full_df))
                    result = build_summary(df_meta(full_df), original_name,
                                           detected_enc2, is_chunked=False)
                    return result
                except Exception as e:
                    _log.warning("小文件全量加载失败，回退采样: %s — %s", file_path, e)
                    # 回退到采样模式

            meta = df_meta(sample_df)
            ctx.loaded_df = None
            ctx.loaded_filename = original_name
            ctx.large_file_path = ""
            ctx.large_file_encoding = ""
            ctx.large_file_suffix = ""
            ctx.source_file_path = file_path          # ★ 记录源文件
            ctx.source_file_suffix = suffix
            _log.info("CSV 采样: %s (%d 采样行, %.1f MB)",
                      original_name, len(sample_df), file_size_mb)

            label = f"已采样（仅前{SAMPLE_ROWS}行）"
            result = build_summary(meta, original_name, detected_enc,
                                   is_chunked=False, label=label)
            result += tail_note
            result += (
                f"\n\n💡 仅探查前 {SAMPLE_ROWS} 行数据结构。"
                f"\n   首次执行 query / clean / export 时将自动全量加载。"
            )

        return result

    # ── query_loaded_data ─────────────────────────────────

    @lc_tool
    def query_loaded_data(expression: str) -> str:
        """
        对已加载的 DataFrame 执行 pandas 表达式，可用变量 df。
        结果超过 50 行自动截断。采样模式下自动触发全量加载。

        常用: df['列'].sum(), df.groupby('列')['列'].sum(),
              df.nlargest(5,'列'), df.corr(), df['列'].pct_change(),
              df.groupby(pd.Grouper(key='日期',freq='M'))['列'].sum()

        Parameters:
        expression: pandas 表达式，如 "df['销售额'].sum()"
        """
        from core.cleaning import _friendly_error, _truncate

        # ★ 采样模式 → 自动全量加载
        if ctx.loaded_df is None and ctx.source_file_path:
            _log.info("查询触发延迟加载")
            if not _ensure_full_loaded(ctx):
                return f"错误: 自动加载失败 — {ctx.source_file_path}"

        if ctx.loaded_df is not None:
            try:
                result = eval(
                    expression,
                    {"df": ctx.loaded_df, "pd": pd, "np": np,
                     "__builtins__": {}},
                )
            except Exception as e:
                return _friendly_error(e, ctx.loaded_df)
            if isinstance(result, pd.DataFrame) and result.shape[0] == result.shape[1] and ".corr(" in expression:
                return "相关性矩阵:\n" + result.to_string()
            source = f"📋 数据来源: 槽位'{ctx.active_slot}' — {ctx.loaded_filename or '内存数据'} ({len(ctx.loaded_df):,}行)\n"
            return source + _truncate(result)

        if ctx.large_file_path:
            if is_aggregation(expression):
                return eval_chunked(expression, ctx.large_file_path,
                                    ctx.large_file_encoding)
            return eval_full_reload(expression, ctx.large_file_path,
                                    ctx.large_file_encoding)

        return "错误: 没有已加载的数据。请先用 read_data_file 读取文件。"

    # ── clean_data ────────────────────────────────────────

    @lc_tool
    def clean_data(expression: str) -> str:
        """
        清洗已加载数据。返回 DataFrame 的 pandas 表达式。
        自动保存快照，可用 undo 撤销。采样模式下自动触发全量加载。

        常用: df.dropna(), df.fillna({'列':0}), df.drop_duplicates(),
              df[df['价格']>0], df.drop(columns=['列']),
              df.rename(columns={'旧':'新'}), df.astype({'列':'int'})

        Parameters:
        expression: 如 "df.dropna().drop_duplicates()"
        """
        # ★ 采样模式 → 自动全量加载
        if ctx.loaded_df is None and ctx.source_file_path:
            _log.info("清洗触发延迟加载")
            if not _ensure_full_loaded(ctx):
                return f"错误: 自动加载失败 — {ctx.source_file_path}"

        if ctx.loaded_df is not None:
            before_rows = len(ctx.loaded_df)
            ctx.snapshot()
            try:
                new_df, summary = clean_in_memory(ctx.loaded_df, expression)
            except (TypeError, ValueError) as e:
                return str(e)
            except Exception as e:
                return f"错误: 清洗失败 — {e}"
            ctx.loaded_df = new_df
            ctx.log_action("clean", expression, before_rows, len(new_df))
            return summary

        if ctx.large_file_path:
            summary, new_path = squash_large(
                ctx.large_file_path, ctx.large_file_encoding,
                expression, ctx.temp_files,
            )
            if new_path:
                ctx.large_file_path = new_path
                ctx.large_file_encoding = "utf-8"
                ctx.large_file_suffix = ".csv"
            return summary

        return "错误: 没有已加载的数据。"

    # ── undo ──────────────────────────────────────────────

    @lc_tool
    def undo() -> str:
        """
        撤销最近一次数据变更（clean_data/merge_file/sql_to_dataframe）。

        Parameters:
        (无)
        """
        return ctx.undo()

    # ── export_data_file ──────────────────────────────────

    @lc_tool
    def export_data_file(output_path: str, expression: str = "") -> str:
        """
        导出数据为 CSV/Excel（由后缀决定），始终保存到 workspace/ 目录。
        可选先执行表达式筛选再导出。采样模式下自动触发全量加载。

        Parameters:
        output_path: 导出路径（相对路径默认保存到 workspace/，如 "report.csv"）
        expression: 可选，导出前先筛选，如 "df.nlargest(50,'销量')" 或空=全量
        """
        resolved, err = _resolve_write_path(output_path, ctx)
        if err:
            return err
        output_path = resolved

        path = Path(output_path)
        suffix = path.suffix.lower()
        if suffix not in (".csv", ".xlsx", ".xls"):
            return f"错误: 不支持的导出格式 {suffix}"

        # ★ 采样模式 → 自动全量加载
        if ctx.loaded_df is None and ctx.source_file_path:
            _log.info("导出触发延迟加载")
            if not _ensure_full_loaded(ctx):
                return f"错误: 自动加载失败 — {ctx.source_file_path}"

        # 获取数据
        if ctx.loaded_df is not None:
            df = ctx.loaded_df
        elif ctx.large_file_path:
            # 大文件模式：警告可能内存不足，仍尝试加载导出
            try:
                import pandas as pd
                # 仅加载前 100,000 行，避免 OOM
                df = pd.read_csv(ctx.large_file_path, nrows=100000)
                _log.warning("大文件导出（仅前10万行）: %s", ctx.large_file_path)
            except Exception as e:
                return f"错误: 读取大文件失败 — {e}"
        else:
            return "错误: 没有已加载的数据。"

        # 可选筛选
        if expression.strip():
            try:
                result = eval(
                    expression,
                    {"df": df, "pd": pd, "np": np, "__builtins__": {}},
                )
                if isinstance(result, pd.DataFrame):
                    df = result
                elif isinstance(result, pd.Series):
                    df = result.to_frame()
                else:
                    return f"错误: 表达式返回了 {type(result).__name__}，需要 DataFrame"
                before_label = f"筛选后: {len(df):,} 行"
            except Exception as e:
                return f"错误: 表达式执行失败 — {e}"
        else:
            before_label = ""

        try:
            if suffix == ".csv":
                df.to_csv(output_path, index=False, encoding="utf-8-sig")
            else:
                df.to_excel(output_path, index=False)
        except Exception as e:
            return f"错误: 导出失败 — {e}"

        size_mb = get_file_size_mb(output_path)
        extra = f"\n{ before_label}" if before_label else ""
        return (
            f"✅ 已导出: {output_path}\n"
            f"📏 {len(df):,} 行 × {len(df.columns)} 列 | 📦 {size_mb:.1f} MB"
            + extra
        )

    # ── merge_file ────────────────────────────────────────

    @lc_tool
    def merge_file(file_path: str = "", on: str = "", how: str = "left",
                   from_slot: str = "") -> str:
        """
        副文件（或另一槽位的 DataFrame）与当前主数据按列 JOIN。自动拍快照，可用 undo 撤销。
        from_slot 非空时直接从命名槽位获取副数据，无需文件路径；否则从 file_path 读取文件。

        Parameters:
        file_path: 副文件路径（from_slot 为空时需要）
        on: 关联列名
        how: 关联方式 (left/inner/right/outer)
        from_slot: 可选，从指定槽位的 DataFrame 作为副数据来源
        """
        main_df = _resolve_df(ctx)
        if main_df is None:
            return "错误: 没有已加载的主数据。"

        side_name = ""
        if from_slot.strip():
            # 从命名槽位获取副数据
            slot_name = from_slot.strip()
            old_active = ctx.active_slot
            ctx.switch(slot_name)
            side_df = _resolve_df(ctx)
            side_name = slot_name
            ctx.switch(old_active)  # 切回原槽位
            if side_df is None:
                return f"错误: 槽位 '{slot_name}' 中没有已加载的数据。"
        else:
            resolved = _resolve_read_path(file_path, ctx)
            if not resolved or resolved.startswith("错误:"):
                return resolved or (
                    f"错误: 文件路径不受支持 — {file_path}\n"
                    f"副文件只能从 raw/ 或 workspace/ 目录读取。"
                )
            file_path = resolved

            path = Path(file_path)
            if not path.exists():
                return f"错误: 文件不存在 — {file_path}"

            suffix = path.suffix.lower()
            if suffix not in (".csv", ".xlsx", ".xls"):
                return f"错误: 不支持的文件格式 {suffix}"

            try:
                if suffix == ".csv":
                    side_df, _ = read_csv_full(str(path))
                else:
                    side_df = pd.read_excel(str(path))
            except Exception as e:
                return f"错误: 读取副文件失败 — {e}"
            side_name = Path(file_path).name

        if on not in main_df.columns:
            return f"错误: 主数据中不存在 '{on}'。可用列: {', '.join(main_df.columns[:12])}"
        if on not in side_df.columns:
            return f"错误: 副数据中不存在 '{on}'。可用列: {', '.join(side_df.columns[:12])}"
        if how not in ("left", "right", "inner", "outer"):
            return f"错误: how 必须为 left/right/inner/outer"

        try:
            merged = main_df.merge(side_df, on=on, how=how, suffixes=('', '_副'))
        except Exception as e:
            return f"错误: 关联失败 — {e}"

        before_rows = len(main_df)
        ctx.snapshot()
        ctx.loaded_df = merged
        ctx.loaded_filename = f"{ctx.loaded_filename} + {side_name}"
        ctx.add_side_file(side_name, side_df)
        ctx.log_action("merge", f"merge({side_name}, on={on}, how={how})",
                       before_rows, len(merged))

        new_cols = [c for c in merged.columns if c not in main_df.columns]
        return (
            f"✅ 关联完成\n"
            f"📏 {before_rows:,} 行 + {len(side_df):,} 行 → {len(merged):,} 行 × {len(merged.columns)} 列\n"
            f"🔗 on='{on}' {how} | 📐 新增: {', '.join(new_cols[:8])}"
            f"{'…' if len(new_cols) > 8 else ''}\n💡 可用 undo 撤销"
        )

    # ── switch_data_slot ──────────────────────────────────

    @lc_tool
    def switch_data_slot(name: str) -> str:
        """
        切换到指定数据槽位（不存在则自动创建）。
        每个槽位独立持有自己的 DataFrame，切换不会丢失数据。
        用于在多个数据源之间切换（如处理7月数据中途查看销售人员表）。

        Parameters:
        name: 槽位名称，如 "7月"、"销售人员"、"6月"
        """
        result = ctx.switch(name)
        # Append slot list for context
        return result + "\n\n" + ctx.list_slots()

    # ── list_data_slots ───────────────────────────────────

    @lc_tool
    def list_data_slots() -> str:
        """
        列出所有数据槽位及其状态（行数、列数、文件名）。
        用于查看当前有哪些数据已加载，以及当前活跃的是哪个槽位。

        Parameters: (无)
        """
        return ctx.list_slots()

    # ── copy_to_workspace ─────────────────────────────────

    @lc_tool
    def copy_to_workspace(source: str, target_name: str = "") -> str:
        """
        将 raw/ 下的文件或文件夹复制到 workspace/。
        如果要修改原始文件，必须先用此工具复制到 workspace 再操作。

        Parameters:
        source: raw/ 下的相对路径，如 "sales.csv" 或 "archive/"
        target_name: 可选新名称；不传则保持原名
        """
        cleaned = source
        for prefix in ("raw/", "raw\\"):
            if cleaned.lower().startswith(prefix.lower()):
                cleaned = cleaned[len(prefix):]
                break
        src = (config.RAW_DIR / cleaned).resolve()
        raw = config.RAW_DIR.resolve()
        ws = Path(ctx.workspace_dir).resolve()
        global_ws = config.WORKSPACE_DIR.resolve()

        if not str(src).startswith(str(raw)):
            return f"错误: 只能从 raw/ 目录复制文件。当前路径: {source}"
        if not src.exists():
            return f"错误: 源文件/文件夹不存在 — {source}"

        dst_name = target_name.strip() if target_name.strip() else src.name
        dst = (ws / dst_name).resolve()
        if not (str(dst).startswith(str(ws)) or str(dst).startswith(str(global_ws))):
            return f"错误: 目标必须在 workspace/ 目录内。"

        try:
            if src.is_dir():
                if dst.exists():
                    shutil.rmtree(dst)
                shutil.copytree(src, dst)
                file_count = sum(1 for _ in dst.rglob("*") if _.is_file())
                return (
                    f"✅ 文件夹已复制到 workspace\n"
                    f"📂 源: raw/{cleaned}\n"
                    f"📂 目标: workspace/{dst.name}\n"
                    f"📏 文件数: {file_count}"
                )
            else:
                dst.parent.mkdir(parents=True, exist_ok=True)
                shutil.copy2(src, dst)
                size_kb = round(dst.stat().st_size / 1024, 1)
                return (
                    f"✅ 文件已复制到 workspace\n"
                    f"📄 源: raw/{cleaned}\n"
                    f"📄 目标: workspace/{dst.name}\n"
                    f"📦 大小: {size_kb:.1f} KB"
                )
        except Exception as e:
            return f"错误: 复制失败 — {e}"

    # ── drop_data_slot ───────────────────────────────────

    @lc_tool
    def drop_data_slot(name: str) -> str:
        """
        删除指定的数据槽位（释放内存）。不能删除当前活跃槽位。

        Parameters:
        name: 要删除的槽位名称
        """
        return ctx.drop_slot(name)

    # ── concat_slots ──────────────────────────────────────

    @lc_tool
    def concat_slots(slot_names: str) -> str:
        """
        纵向拼接多个槽位的 DataFrame（UNION ALL），结果存入当前槽位。
        用于合并多个月份/来源的数据。自动拍快照，可用 undo 撤销。

        Parameters:
        slot_names: 逗号分隔的槽位名称，如 "4-5月, 6月, 7月"
        """
        names = [n.strip() for n in slot_names.split(",") if n.strip()]
        if len(names) < 2:
            return "错误: 至少需要 2 个槽位名称（逗号分隔）"

        dfs = []
        old_active = ctx.active_slot
        for name in names:
            if name not in ctx._slots:
                return f"错误: 槽位 '{name}' 不存在"
            ctx.switch(name)
            df = ctx._slots[name].df
            if df is None:
                # Try to resolve sampled data
                df = _resolve_df(ctx)
            if df is None:
                ctx.switch(old_active)
                return f"错误: 槽位 '{name}' 中没有已加载的数据"
            dfs.append((name, df))

        ctx.switch(old_active)
        try:
            merged = pd.concat([d for _, d in dfs], ignore_index=True)
        except Exception as e:
            return f"错误: 拼接失败 — {e}"

        before = len(ctx.loaded_df) if ctx.loaded_df is not None else 0
        ctx.snapshot()
        ctx.loaded_df = merged
        ctx.loaded_filename = " + ".join(names)
        ctx.log_action("concat", slot_names, before, len(merged))

        parts = [f"{n}: {len(d):,} 行" for n, d in dfs]
        return (
            f"✅ 纵向拼接完成\n"
            f"📋 来源: {', '.join(n for n, _ in dfs)}\n"
            f"📏 {' → '.join(parts)} → 总计 {len(merged):,} 行 × {len(merged.columns)} 列\n"
            f"💡 可用 undo 撤销"
        )

    return [
        list_files, read_data_file, query_loaded_data, clean_data, undo,
        export_data_file, merge_file, switch_data_slot, list_data_slots,
        copy_to_workspace, drop_data_slot, concat_slots,
    ]
